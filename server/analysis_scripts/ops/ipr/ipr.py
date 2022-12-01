# -*- coding: utf-8 -*-
"""
Created on Wed Jun 24 11:08:27 2020

@author: Meryll
"""

from datetime import datetime, timedelta
import openpyxl
import os
import pandas as pd
import sys

import bulletin
import dtr
import lib as ipr_lib
import raininfo
import shifteval
import sms

import aim
import eosr
import fyipermission
import gndmeas
import measreminder
import webrelease

sys.path.append(os.path.dirname(os.path.realpath(__file__)))
import ops.lib.lib as lib

output_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..//input_output//'))


def format_cells():
    wbook=openpyxl.load_workbook(output_path + 'monitoring_ipr.xlsx')
    for sheet_name in pd.read_excel(output_path + 'monitoring_ipr.xlsx', sheet_name=None).keys():
        sheet=wbook[sheet_name]
        sheet.merge_cells('A2:E2')
        sheet.merge_cells('C3:E3')
        sheet.merge_cells('A4:A11')
        sheet.merge_cells('B4:B11')
        sheet.merge_cells('C4:C9')
        sheet.merge_cells('C10:C11')
        sheet.merge_cells('C12:E12')
        sheet.merge_cells('A13:A20')
        sheet.merge_cells('C13:E13')
        sheet.merge_cells('C14:E14')
        sheet.merge_cells('C15:E15')
        sheet.merge_cells('C16:E16')
        sheet.merge_cells('C17:E17')
        sheet.merge_cells('C18:E18')
        sheet.merge_cells('C19:E19')
        sheet.merge_cells('C20:E20')
        sheet.merge_cells('A21:A32')
        sheet.merge_cells('B23:B29')
        sheet.merge_cells('C21:E21')
        sheet.merge_cells('C22:E22')
        sheet.merge_cells('C23:C29')
        sheet.merge_cells('C30:E30')
        sheet.merge_cells('C31:E31')
        sheet.merge_cells('C32:E32')
        sheet.merge_cells('A33:B40')
        sheet.merge_cells('C33:C38')
        sheet.merge_cells('C39:C40')    
    wbook.save(output_path + 'monitoring_ipr.xlsx')
    

def timeline(year, quarter):
    if quarter == 4:
        end = "{}-01-01".format(year+1)
    else:
        end = "{}-{}-01".format(year, str(3*quarter + 1).zfill(2))
    end = pd.to_datetime(end)
    start = pd.to_datetime("{}-{}-01".format(year, str(3*quarter - 2).zfill(2)))
    return start, end


def get_shift(key, sheet_name):
    df = ipr_lib.get_sheet(key, sheet_name)
    df = df.drop([col for col in df.columns if col.startswith('Unnamed')], axis=1)
    df.loc[:, 'Date'] = pd.to_datetime(df.loc[:, 'Date'].ffill())
    df.loc[:, 'Shift'] = pd.to_timedelta(df.loc[:, 'Shift'].map({'AM': timedelta(hours=8), 'PM': timedelta(hours=20)}))
    df.loc[:, 'ts'] = pd.to_datetime(df.loc[:, 'Date'] + df.loc[:, 'Shift'])
    return df.loc[:, ['ts', 'IOMP-MT', 'IOMP-CT']]


def main(start, end, update_existing=True, update_dtr=True, recompute=True, mysql=True, to_csv=False):
    key = "1UylXLwDv1W1ukT4YNoUGgHCHF-W8e3F8-pIg1E024ho"
    date_range = pd.date_range(start=start, end=end, freq='M', closed='left')
    shift_sched = pd.DataFrame()
    for ts in date_range:
        sheet_name = ts.strftime('%B %Y')
        shift_sched = shift_sched.append(get_shift(key, sheet_name))
    shift_sched = shift_sched.loc[(shift_sched.ts > start) & (shift_sched.ts < end)]
    
    if update_existing:
        monitoring_ipr = pd.read_excel(output_path + 'monitoring_ipr.xlsx', sheet_name=None)
        writer = pd.ExcelWriter(output_path + 'monitoring_ipr.xlsx')
        for name in monitoring_ipr.keys() -set({'Summary'}):
            indiv_ipr = monitoring_ipr[name]
            indiv_ipr = indiv_ipr.loc[:, ~indiv_ipr.columns.str.contains('Unnamed:', na=False)]
            for shift_type in ['MT', 'CT']:
                curr_shift = shift_sched.loc[shift_sched['IOMP-{}'.format(shift_type)] == name, :]
                shift_list = set(curr_shift.ts) - set(map(pd.to_datetime, indiv_ipr.columns[5:]))
                for ts in sorted(shift_list):
                    indiv_ipr.loc[indiv_ipr.Category == 'Shift', str(ts)] = shift_type
            monitoring_ipr[name] = indiv_ipr

    else:
        writer = pd.ExcelWriter(output_path + 'monitoring_ipr.xlsx')
        grading_system = pd.read_excel('baseline.xlsx', sheet_name='format')
        monitoring_ipr = {}
    
        personnel_sheet = "personnel"    
        
        name_df = ipr_lib.get_sheet(key, personnel_sheet)
        name_list = name_df.loc[name_df.current == 1, 'Nickname'].values
            
        for name in name_list:
            indiv_ipr = grading_system.copy()
            for shift_type in ['MT', 'CT']:
                curr_shift = shift_sched.loc[shift_sched['IOMP-{}'.format(shift_type)] == name, :]
                for ts in curr_shift.ts:
                    indiv_ipr.loc[indiv_ipr.Category == 'Shift', str(ts)] = shift_type
            monitoring_ipr[name] = indiv_ipr
    
    for sheet_name, xlsxdf in monitoring_ipr.items():
        xlsxdf.to_excel(writer, sheet_name, index=False)
    writer.save()
    
    site_names = lib.get_site_names()
    sched = lib.release_sched(start, end, mysql=mysql, to_csv=to_csv)
    eval_df = shifteval.get_eval_df()

    sms.main(start, end, sched, site_names, eval_df, mysql=True, to_csv=to_csv)
    bulletin.main(start, end, sched, site_names, eval_df, mysql=mysql, to_csv=to_csv)
#    dtr.main(sched, update_dtr)
    raininfo.main(start, end, sched, site_names, eval_df, mysql=mysql, to_csv=to_csv)

    measreminder.main(start, end, sched, mysql=mysql) #fix late sending of gndmeas
    
    gndmeas.main(start, end, sched, eval_df, mysql=mysql)
    webrelease.main(start, end, sched, eval_df, mysql=mysql)
    eosr.main(start, end, sched, eval_df, mysql=mysql)
    shifteval.eval_timeliness(sched, eval_df)
    fyipermission.main(start, end, sched, mysql=mysql)
    aim.main(start, end, sched, eval_df, mysql=mysql)
    
    format_cells()
    
###############################################################################
if __name__ == "__main__":
    run_start = datetime.now()
    
    start = pd.to_datetime('2021-12-01')
    end = pd.to_datetime('2022-06-01 08:00')
    main(start, end, update_dtr=True, mysql=True, recompute=True, update_existing=True, to_csv=True)
    
    runtime = datetime.now() - run_start
    print("runtime = {}".format(runtime))